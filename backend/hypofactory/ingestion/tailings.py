"""Парсер отчёта института по хвостам (Excel) в структурированную модель потерь.

Отчёт описывает баланс отвальных хвостов обогатительной фабрики: сколько ценного
металла (Элемент 28 = Ni, Элемент 29 = Cu) теряется, с разбивкой по классам
крупности и по минеральным формам, а также деление на извлекаемый /
неизвлекаемый металл. Ключевое различие для последующей диагностики:
  - «Закрытый Pnt/Cp» — металл заперт в сростках → проблема раскрытия (измельчение);
  - «Раскрытый Pnt/Cp» — металл вскрыт, но потерян → проблема флотации.

Парсер устойчив к вариациям формата (наличие/отсутствие пирротиновых хвостов,
блоки «Факт»/«Расчёт», ячейки #REF!, разные подписи классов вроде «+71» и
«-125 +71»). Привязка идёт к текстовым меткам, а не к номерам строк.
"""
from __future__ import annotations

from dataclasses import dataclass, field, asdict
from typing import Any, Optional

import openpyxl

# Канонический порядок классов крупности (мкм), от крупного к тонкому.
CLASS_ORDER = ["+125", "-125+71", "-71+45", "-45+20", "-20+10", "-10"]

# Нормализация подписи класса (после удаления пробелов) к каноническому виду.
_CLASS_VARIANTS = {
    "+125": "+125",
    "+71": "-125+71",
    "-125+71": "-125+71",
    "-71+45": "-71+45",
    "-45+20": "-45+20",
    "-20+10": "-20+10",
    "-10": "-10",
}


def _norm_class(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    t = str(raw).lower().replace("мкм", "").replace(" ", "").strip()
    return _CLASS_VARIANTS.get(t)


def _num(v: Any) -> Optional[float]:
    """Приводит ячейку к float; ошибки Excel (#REF!, #DIV/0!) и пустоты → None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if not s or s.startswith("#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _classify_form(label: str) -> tuple[str, str, bool, bool]:
    """Возвращает (канон. имя, категория, извлекаем Ni, извлекаем Cu).

    Категории: 'open' (раскрытый, извлекаем), 'locked' (закрытый, извлекаем но
    заперт), 'norec' (физически неизвлекаемо текущей технологией).
    """
    s = label.lower()
    if "раскрыт" in s:
        return "Раскрытый Pnt/Cp", "open", True, True
    if "закрыт" in s:
        return "Закрытый Pnt/Cp", "locked", True, True
    if "миллерит" in s:
        return "Миллерит", "open", True, False  # носитель только Ni
    if "примес" in s or "пирротин" in s:
        return "Примесь в пирротине", "norec", False, False
    if "силикат" in s or "валлериит" in s:
        return "Силикатная форма/Валлериит", "norec", False, False
    if "пирит" in s:
        return "Пирит/Другие сульфиды", "norec", False, False
    return label.strip(), "other", False, False


@dataclass
class MineralLoss:
    form: str
    category: str          # open | locked | norec | other
    recoverable_ni: bool
    recoverable_cu: bool
    loss_pct_28: Optional[float]   # доля потерь металла 28 внутри класса, %
    loss_t_28: Optional[float]     # тонны металла 28
    loss_pct_29: Optional[float]
    loss_t_29: Optional[float]


@dataclass
class SizeClass:
    name: str                      # канонический класс
    raw_name: str
    share_pct: Optional[float]     # доля класса в массе хвостов, %
    loss_t_28: Optional[float]     # всего металла 28 в классе, т
    loss_t_29: Optional[float]
    minerals: list[MineralLoss] = field(default_factory=list)
    recoverable_t_28: Optional[float] = None
    nonrecoverable_t_28: Optional[float] = None
    recoverable_t_29: Optional[float] = None
    nonrecoverable_t_29: Optional[float] = None


@dataclass
class TailingsStream:
    kind: str                      # породные | пирротиновые
    smt: Optional[float]
    loss_t_28: Optional[float]
    loss_t_29: Optional[float]
    classes: list[SizeClass] = field(default_factory=list)


@dataclass
class TailingsReport:
    source_file: str
    feed_smt: Optional[float] = None
    feed_grade_28: Optional[float] = None
    feed_grade_29: Optional[float] = None
    tails_smt: Optional[float] = None
    tails_grade_28: Optional[float] = None
    tails_grade_29: Optional[float] = None
    tails_loss_t_28: Optional[float] = None
    tails_loss_t_29: Optional[float] = None
    streams: list[TailingsStream] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return asdict(self)


def _cell(row: list, idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def _label(row: list) -> str:
    v = _cell(row, 1)
    return str(v).strip() if v is not None else ""


def parse_tailings_report(path: str, sheet: str = "Итог") -> TailingsReport:
    """Разбирает баланс хвостов. Устойчив к вариациям структуры книги:
    приоритет у листа `sheet`, но если данных на нём нет — перебираются ВСЕ
    листы, берётся тот, где распознано больше всего (потоки × классы)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    names = ([sheet] if sheet in wb.sheetnames else []) \
        + [n for n in wb.sheetnames if n != sheet]

    best: Optional[TailingsReport] = None
    best_name, best_score = "", -1
    for name in names:
        rows = [[c.value for c in r] for r in wb[name].iter_rows()]
        rep = _parse_rows(rows, source_file=path)
        score = sum(len(s.classes) for s in rep.streams) * 10 \
            + len(rep.streams) + (1 if rep.tails_smt is not None else 0)
        if score > best_score:
            best, best_name, best_score = rep, name, score
        if best_score > 0 and name == sheet:
            break  # предпочтительный лист дал данные — дальше не ищем

    assert best is not None
    if best_score <= 0:
        raise ValueError(
            "Не удалось распознать баланс хвостов ни на одном листе книги: "
            "не найдены маркеры («Хвосты …», «Класс крупности», классы, «доля потерь»). "
            f"Листы в файле: {', '.join(wb.sheetnames)}. Ожидается отчёт института по хвостам.")
    if best_name != sheet:
        best.warnings.insert(0, f"Данные распознаны на листе «{best_name}» (лист «{sheet}» не найден или пуст).")
    _validate(best)
    return best


def _parse_rows(rows: list[list], source_file: str) -> TailingsReport:
    n = len(rows)
    report = TailingsReport(source_file=source_file)
    report.streams = []

    current_stream: Optional[TailingsStream] = None
    class_index: dict[str, SizeClass] = {}
    # «Ожидающие» параметры потока, считанные со строк «Хвосты X ...» до таблицы
    pend_kind: Optional[str] = None
    pend_smt: Optional[float] = None
    pend_l28: Optional[float] = None
    pend_l29: Optional[float] = None
    i = 0
    feed_captured = False

    while i < n:
        row = rows[i]
        lbl = _label(row)
        low = lbl.lower()

        # --- Верхнеуровневые метрики ---
        if low.startswith("итого") and not feed_captured and not report.streams \
                and _num(_cell(row, 2)) is not None:
            report.feed_smt = _num(_cell(row, 2))
            report.feed_grade_28 = _num(_cell(row, 3))
            report.feed_grade_29 = _num(_cell(row, 5))
            feed_captured = True
        elif low.startswith("отвальные хвосты") and _num(_cell(row, 2)) is not None:
            if report.tails_smt is None:
                report.tails_smt = _num(_cell(row, 2))
                report.tails_grade_28 = _num(_cell(row, 3))
                report.tails_loss_t_28 = _num(_cell(row, 4))
                report.tails_grade_29 = _num(_cell(row, 5))
                report.tails_loss_t_29 = _num(_cell(row, 6))

        # --- Строка-заголовок потока (может встречаться несколько раз) ---
        if low.startswith("хвосты ") and _num(_cell(row, 2)) is not None:
            if "пирротин" in low:
                pend_kind = "пирротиновые"
            elif "породн" in low:
                pend_kind = "породные"
            elif "отвальн" in low or "общ" in low:
                pend_kind = "отвальные (общие)"
            else:
                pend_kind = lbl
            pend_smt = _num(_cell(row, 2))
            l28, l29 = _num(_cell(row, 4)), _num(_cell(row, 6))
            if l28 is not None:
                pend_l28 = l28
            if l29 is not None:
                pend_l29 = l29

        # --- Таблица распределения по классам: здесь материализуем поток ---
        if lbl.startswith("Класс крупности"):
            current_stream = TailingsStream(
                kind=pend_kind or "породные",
                smt=pend_smt if pend_smt is not None else report.tails_smt,
                loss_t_28=pend_l28,
                loss_t_29=pend_l29,
            )
            report.streams.append(current_stream)
            class_index = {}
            pend_l28 = pend_l29 = None  # следующий поток задаст свои значения
            j = i + 1
            while j < n:
                r2 = rows[j]
                l2 = _label(r2)
                l2low = l2.lower()
                if l2low.startswith("итого"):
                    break
                # начались минералогические блоки (по колонке «Доля потерь»)
                if "доля потерь" in str(_cell(r2, 3) or "").lower():
                    break
                if l2low.startswith("хвосты") or l2.startswith("Класс крупности"):
                    break
                if l2 == "":
                    j += 1
                    continue
                canon = _norm_class(l2)
                if canon is None:
                    j += 1
                    continue
                sc = SizeClass(
                    name=canon,
                    raw_name=l2,
                    share_pct=_num(_cell(r2, 2)),
                    loss_t_28=_num(_cell(r2, 4)),
                    loss_t_29=_num(_cell(r2, 6)),
                )
                current_stream.classes.append(sc)
                class_index[canon] = sc
                j += 1
            i = j
            continue

        # --- Минералогический блок конкретного класса ---
        # Определяем по колонке «Доля потерь…» (метка класса бывает «+71» без «мкм»).
        if "доля потерь" in str(_cell(row, 3) or "").lower() and _norm_class(lbl) is not None:
            canon = _norm_class(lbl)
            sc = class_index.get(canon) if canon else None
            j = i + 1
            while j < n:
                r2 = rows[j]
                l2 = _label(r2)
                l2low = l2.lower()
                if l2low.startswith("итого (проверка)"):
                    j += 1
                    continue
                if l2low.startswith("извлекаемый металл"):
                    if sc is not None:
                        sc.recoverable_t_28 = _num(_cell(r2, 4))
                        sc.recoverable_t_29 = _num(_cell(r2, 6))
                    j += 1
                    continue
                if l2low.startswith("не извлекаемый металл"):
                    if sc is not None:
                        sc.nonrecoverable_t_28 = _num(_cell(r2, 4))
                        sc.nonrecoverable_t_29 = _num(_cell(r2, 6))
                    j += 1
                    break  # блок класса закончился
                # следующий блок/секция
                if _norm_class(l2) is not None or l2low.startswith("хвосты") \
                        or l2.startswith("Класс крупности") \
                        or "доля потерь" in str(_cell(r2, 3) or "").lower():
                    break
                if l2 == "" or l2low.startswith(("потери", "свободный слот")):
                    j += 1
                    continue
                form, cat, rec_ni, rec_cu = _classify_form(l2)
                if sc is not None:
                    sc.minerals.append(MineralLoss(
                        form=form, category=cat,
                        recoverable_ni=rec_ni, recoverable_cu=rec_cu,
                        loss_pct_28=_num(_cell(r2, 3)), loss_t_28=_num(_cell(r2, 4)),
                        loss_pct_29=_num(_cell(r2, 5)), loss_t_29=_num(_cell(r2, 6)),
                    ))
                j += 1
            i = j
            continue

        i += 1

    return report


def _validate(report: TailingsReport) -> None:
    if not report.streams:
        report.warnings.append("Не найдено ни одного потока хвостов.")
    for st in report.streams:
        seen = {c.name for c in st.classes}
        missing = [c for c in CLASS_ORDER if c not in seen]
        if missing:
            report.warnings.append(
                f"Поток '{st.kind}': нет классов {missing} (возможно нулевая доля)."
            )
        for c in st.classes:
            if not c.minerals and (c.share_pct or 0) > 0:
                report.warnings.append(
                    f"Поток '{st.kind}', класс {c.name}: нет минералогии (#REF!?)."
                )


if __name__ == "__main__":
    import sys
    rep = parse_tailings_report(sys.argv[1])
    print(rep.to_dict())
